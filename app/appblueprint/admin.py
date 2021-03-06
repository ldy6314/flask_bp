from flask import Blueprint, send_file, session, jsonify, flash, current_app
from forms import AddForm, AddSubjectForm, UploadClassForm, AddStudentForm, UploadSubjectsForm
from flask import render_template
from models import Subject
from extensions import db
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Side
from io import BytesIO
from flask_login import login_required
from utils import redirect_back, random_filename, get_subjects
import os


admin_bp = Blueprint('admin', __name__)


@admin_bp.before_request
@login_required
def login_required():
    pass


@admin_bp.route('/')
def index():
    return "I am appblueprint index"


@admin_bp.route('/add', methods=['GET', 'POST'])
def add():
    form = AddForm()
    if form.validate_on_submit():
        for i in form.data:
            print(i, form.data[i])

    return render_template('add.html', form=form)


@admin_bp.route('/add_subject', methods=['GET', 'POST'])
def add_subject():
    form = AddSubjectForm()
    if form.validate_on_submit():
        subject = Subject(name=form.data['name'], time=form.data['time'], price=form.data['price'],
                          remark=form.data['remark'])

        db.session.add(subject)
        db.session.commit()
        res = Subject.query.all()
        for i in res:
            print(i.name)
        return jsonify({'result': "插入成功", "type": "alert alert-success"})
    else:
        return jsonify({'result': "项目已经存在无法添加", "type": "alert  alert-danger"})


@admin_bp.route('/download_subjects')
def download_subjects():
    wb = Workbook()
    ws = wb.worksheets[0]
    ws.append(["项目开设表", "", "", "", ""])
    ws.append(["编号", "名称", "时间", "价格", "备注"])
    res = Subject.query.all()
    for i in enumerate(res, 1):
        print(i[0], i[1].name, i[1].time, i[1].price, i[1].remark)
        ws.append([i[0], i[1].name, i[1].time, i[1].price, i[1].remark])
    ws.merge_cells("A1:E1")
    border = Border(
        left=Side(style='medium', color='FF000000'),
        right=Side(style='medium', color='FF000000'),
        bottom=Side(style='medium', color='FF000000'),
        top=Side(style='medium', color='FF000000')
    )
    align_center = Alignment(horizontal='center', vertical='center')
    ws_area = ws["A1:E22"]
    for row in ws_area:
        for cell in row:
            cell.alignment = align_center
            cell.border = border
    for i in "BCE":
        ws.column_dimensions[i].width = 30

    virtual_book = BytesIO()
    wb.save(virtual_book)
    virtual_book.seek(0)
    rv = send_file(virtual_book, as_attachment=True, attachment_filename="test.xlsx")
    rv.headers['Content-Disposition'] += ";filename*=utf-8' 'test.xlsx"
    return rv


@admin_bp.route('/school_admin', methods=['GET', 'POST'])
def school_admin():
    permission = session['permission']
    print(permission)
    if permission == 2:
        form = UploadSubjectsForm()
        subjects = Subject.query.all()
        form1 = AddSubjectForm()
        return render_template('schooladmin.html', form=form, form1=form1, subjects=subjects)
    else:
        return render_template("permission_deny.html")


@admin_bp.route('/class_admin')
def class_admin():
    form = UploadClassForm()
    form1 = AddStudentForm()
    permission = session['permission']
    if permission == 1:
        return render_template('classadmin.html', form=form, form1=form1)
    return render_template("permission_deny.html")


@admin_bp.route('/set_canceled/<int:subject_id>')
def set_canceled(subject_id):
    subject = Subject.query.filter_by(id=subject_id).first()
    subject.canceled = 1
    db.session.commit()
    return redirect_back('/')


@admin_bp.route('/delete/<int:subject_id>', methods=['GET', 'POST'])
def delete_subject(subject_id):
    subject = Subject.query.filter_by(id=subject_id).first()
    db.session.delete(subject)
    flash("已经删除 {}".format(subject.name))
    db.session.commit()
    return redirect_back('/')


@admin_bp.route('/upload_subjects', methods=['GET', 'POST'])
def upload_subjects():
    path = os.path.join(current_app.root_path, 'upload\\')
    print("path=", path)
    form = UploadSubjectsForm()
    if form.validate_on_submit():
        f = form.file.data
        filename = random_filename(f.filename)
        f.save(path+filename)
        subjects = get_subjects(path+filename)
        infos = []
        for i in subjects:
            subject = Subject(name=i[0], time=i[1], price=int(i[2].strip('元')), canceled=0, remark="")
            infos.append(subject)

        for i in infos:
            try:
                db.session.add(i)
                db.session.commit()
            except Exception as e:
                print(e)


    else:
        flash('上传失败')
    
    return "hehe"